import json
import os
import shutil
import zipfile
import time
import random
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys 
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
from selenium.webdriver.common.action_chains import ActionChains


# File to store accounts and proxy data
EXCEL_FILE = "accounts_and_proxies.xlsx"


# Helper to fetch a new proxy from proxy2.webshare.io API
def get_proxy(api_key, page=1, page_size=100):
    url = f"https://proxy.webshare.io/api/v2/proxy/list/?mode=direct&page={page}&page_size={page_size}"
    response = requests.get(
        url,
        headers={"Authorization": f"Token {api_key}"}
    )
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception("Failed to fetch proxy: ", response.text)

def load_saved_proxies(file_name="valid_proxies.json"):
    """Load saved proxies from a file if it exists."""
    if os.path.exists(file_name):
        with open(file_name, "r") as file:
            return json.load(file)
    return None

def save_proxies_to_file(proxies, file_name="valid_proxies.json"):
    """Save proxies to a file."""
    with open(file_name, "w") as file:
        json.dump(proxies, file, indent=4)

def fetch_all_proxies(api_key):
    """Fetch all proxies from Webshare.io."""
    all_proxies = []
    page = 1
    while True:
        data = get_proxy(api_key, page=page)
        proxies = data.get("results", [])
        all_proxies.extend(proxies)
        next_page = data.get("next")
        if not next_page:
            break
        page += 1
    return all_proxies

def get_or_fetch_proxies(api_key, file_name="valid_proxies.json"):
    """Load proxies from file if saved, else fetch and save them."""
    saved_proxies = load_saved_proxies(file_name)
    if saved_proxies:
        print(f"Loaded {len(saved_proxies)} proxies from {file_name}.")
        return saved_proxies
    else:
        print("No saved proxies found. Fetching new proxies...")
        proxies = fetch_all_proxies(api_key)
        valid_proxies = [
            {
                "id": proxy["id"],
                "username": proxy["username"],
                "password": proxy["password"],
                "proxy_address": proxy["proxy_address"],
                "port": proxy["port"],
                "country_code": proxy["country_code"],
                "city_name": proxy["city_name"]
            }
            for proxy in proxies if proxy["valid"]
        ]
        save_proxies_to_file(valid_proxies, file_name)
        print(f"Fetched and saved {len(valid_proxies)} valid proxies.")
        return valid_proxies
    
    
# Filter proxies that have not been used
def get_unused_proxy(proxies, used_proxies):
    for proxy in proxies:
        proxy_id = proxy["id"]
        if proxy_id not in used_proxies:
            return proxy
    raise Exception("No unused proxies available.")

# Random tech-themed details generator
def generate_account_details():
    first_names = ["Alex", "Chris", "Jordan", "Taylor", "Morgan", "Sam"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Taylor", "Anderson"]
    domains = ["gmail.com", "yahoo.com", "outlook.com"]
    
    first_name = random.choice(first_names)
    last_name = random.choice(last_names)
    username = f"{first_name.lower()}.{last_name.lower()}{random.randint(100, 999)}"
    email = f"{username}@{random.choice(domains)}"
    password = f"Passw0rd!{random.randint(1000, 9999)}"
    
    return {
        "first_name": first_name,
        "last_name": last_name,
        "username": username,
        "email": email,
        "password": password
    }


# Save account details and proxy to an Excel file
def save_to_excel(account_details, proxy, file_name=EXCEL_FILE):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["First Name", "Last Name", "Username", "Email", "Password", "Proxy ID", "Proxy Address", "Proxy Username", "Proxy Password", "Port"])
    
    sheet.append([
        account_details["first_name"],
        account_details["last_name"],
        account_details["username"],
        account_details["email"],
        account_details["password"],
        proxy["id"],
        proxy["proxy_address"],
        proxy["username"],
        proxy["password"],
        proxy["port"]
    ])
    
    workbook.save(file_name)


# Load used proxy IDs from Excel
def load_used_proxies(file_name=EXCEL_FILE):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active 
        used_proxies = {row[5] for row in sheet.iter_rows(min_row=2, values_only=True)}
        return used_proxies
    except FileNotFoundError:
        return set()

extension_folder = "chrome_proxy_extension"
zip_file_path = "chrome_proxy_extension.zip"
def get_zip_file_path(proxy_host, proxy_port, proxy_username, proxy_password):
    cleanup()
    create_background_js(proxy_host, proxy_port, proxy_username, proxy_password)
    add_manifest()
    create_zip_from_directory(extension_folder, zip_file_path)
    return zip_file_path

def create_background_js(host, port, username, password):
    background_js_content = f"""
var config = {{
    mode: 'fixed_servers',
    rules: {{
        singleProxy: {{
            scheme: 'http',
            host: '{host}',
            port: parseInt('{port}')
        }},
        bypassList: ['localhost']
    }}
}};

chrome.proxy.settings.set({{ value: config, scope: 'regular' }}, function () {{}});

function callbackFn(details) {{
    return {{
        authCredentials: {{
            username: '{username}',
            password: '{password}'
        }}
    }};
}}

chrome.webRequest.onAuthRequired.addListener(
    callbackFn,
    {{ urls: ['<all_urls>'] }},
    ['blocking']
);
"""
    create_directory_and_write_file(extension_folder, "background.js", background_js_content)
    add_manifest()

def add_manifest():
    manifest_json_content = """
{
    "manifest_version": 2,
    "name": "Chrome Proxy Extension",
    "version": "1.0",
    "permissions": [
        "proxy",
        "tabs",
        "unlimitedStorage",
        "storage",
        "<all_urls>",
        "webRequest",
        "webRequestBlocking"
    ],
    "background": {
        "scripts": ["background.js"],
        "persistent": true
    },
    "minimum_chrome_version": "22.0.0"
}
"""
    create_directory_and_write_file(extension_folder, "manifest.json", manifest_json_content)

def create_directory_and_write_file(folder_path, file_name, content):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    file_path = os.path.join(folder_path, file_name)
    with open(file_path, "w") as file:
        file.write(content)

def cleanup():
    try:
        if os.path.exists(zip_file_path):
            os.remove(zip_file_path)
        if os.path.exists(extension_folder):
            shutil.rmtree(extension_folder)
    except Exception as e:
        print(f"Error during cleanup: {e}")

def create_zip_from_directory(source_folder, zip_file):
    with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(source_folder):
            for file in files:
                zipf.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), source_folder))

# Automate Instagram registration
def register_account(api_key):
    proxies = get_or_fetch_proxies(api_key)
    used_proxies = load_used_proxies()  
    
    proxy = get_unused_proxy(proxies, used_proxies) 
    zip_path = get_zip_file_path(proxy['proxy_address'], proxy['port'], proxy['username'], proxy['password']) 
    
    chrome_options = Options()
    #chrome_options.add_argument("--headless")  # Run browser in headless mode 
    chrome_options.add_extension(zip_path);   
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage") 
    
    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get("https://www.instagram.com/accounts/emailsignup/")  
        time.sleep(5 + random.random() * 5)  # Add delay to mimic human behavior
        
        account_details = generate_account_details()
        
        # Fill out the form
        driver.find_element(By.NAME, "emailOrPhone").send_keys(account_details["email"])
        time.sleep(1 + random.random() * 3)
        driver.find_element(By.NAME, "fullName").send_keys(f"{account_details['first_name']} {account_details['last_name']}")
        time.sleep(1 + random.random() * 3)
        driver.find_element(By.NAME, "username").send_keys(account_details["username"])
        time.sleep(1 + random.random() * 3)
        driver.find_element(By.NAME, "password").send_keys(account_details["password"])
        time.sleep(1 + random.random() * 3)
         
        # Submit the form
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
        #ActionChains(driver).send_keys(Keys.RETURN).perform()
        time.sleep(5 + random.random() * 5)
        
        # Save details
        save_to_excel(account_details, proxy)
        print("Account registered successfully:", account_details)
    
    except Exception as e:
        print("Error during registration:", e)
    
    finally:
        #driver.quit()
        time.sleep(21000)


if __name__ == "__main__":
    API_KEY = "jp851bkqmezcse9h39a26rn2bvv28yxn4sm4tgcs"
    for _ in range(3):  # Adjust the number of accounts to create
        register_account(API_KEY)
